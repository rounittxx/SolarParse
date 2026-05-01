# Loads the template, writes ONLY into mapped input cells, leaves
# everything else (formulas, formatting, charts) untouched.
#
# The big rule: keep_vba=True is irrelevant because the template is xlsx,
# but we set keep_links=True and we DO NOT pass data_only=True -- that
# would strip formulas the moment we save.

import copy
from pathlib import Path

from openpyxl import load_workbook

from src.config import FIELD_TO_CELL, INPUT_SHEET


def _looks_like_formula(value) -> bool:
    return isinstance(value, str) and value.startswith("=")


def fill_template(template_path: str | Path, fields: dict, output_path: str | Path) -> Path:
    """
    fields: {field_key: value} -- comes straight from the LLM/UI form.
    Returns the path to the new file.
    """
    template_path = Path(template_path)
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if not template_path.exists():
        raise FileNotFoundError(f"template missing: {template_path}")

    # NOTE: data_only=False is the default, but spelling it out so it's
    # obvious this is intentional. We need formulas preserved.
    wb = load_workbook(template_path, data_only=False)

    if INPUT_SHEET not in wb.sheetnames:
        raise ValueError(
            f"template has no '{INPUT_SHEET}' sheet. "
            f"Found: {wb.sheetnames}. Update INPUT_SHEET in src/config.py."
        )
    ws = wb[INPUT_SHEET]

    written, skipped = [], []
    for key, cell_ref in FIELD_TO_CELL.items():
        if key not in fields:
            continue
        value = fields[key]
        if value is None or value == "":
            continue

        cell = ws[cell_ref]
        # safety net -- if someone put a formula in a cell we were
        # told is an input, refuse to clobber it.
        if _looks_like_formula(cell.value):
            skipped.append((key, cell_ref, "would overwrite formula"))
            continue

        cell.value = value
        written.append((key, cell_ref, value))

    wb.save(output_path)

    # tiny audit log -- handy when something didn't end up where we expected
    fill_template.last_report = {     # type: ignore[attr-defined]
        "written": written,
        "skipped": skipped,
        "output": str(output_path),
    }

    return output_path


def report():
    return getattr(fill_template, "last_report", None)
