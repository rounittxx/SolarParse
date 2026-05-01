# Generates a clean default template that mimics how solar load
# calculators are usually laid out:
#   Sheet 1 "Inputs"  -- bill data (this is what we fill)
#   Sheet 2 "Solar"   -- live formulas referencing the inputs
#
# When Energybae shares their real template, drop it into templates/
# and skip this file -- the filler.py will use whatever sits there.

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.config import FIELD_TO_CELL, FIELDS, INPUT_SHEET


# colours roughly matching Energybae's brand (green + dark)
BRAND_GREEN = "1FA463"
DARK = "1F2937"
SOFT_GREY = "F3F4F6"
INPUT_FILL = "FFFBEA"   # pale yellow so the user knows what's editable

THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _header(cell, text, fill=BRAND_GREEN, color="FFFFFF"):
    cell.value = text
    cell.font = Font(bold=True, color=color, size=12)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _label(cell, text):
    cell.value = text
    cell.font = Font(bold=True, color=DARK)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _input(cell, placeholder=""):
    cell.value = placeholder
    cell.fill = PatternFill("solid", fgColor=INPUT_FILL)
    cell.border = BORDER
    cell.alignment = Alignment(horizontal="left", vertical="center")


def build(path: str | Path) -> Path:
    wb = Workbook()

    # ---- Sheet 1: Inputs --------------------------------------------------
    s1 = wb.active
    s1.title = INPUT_SHEET
    s1.column_dimensions["A"].width = 2
    s1.column_dimensions["B"].width = 28
    s1.column_dimensions["C"].width = 38

    s1.merge_cells("B2:C2")
    _header(s1["B2"], "Energybae -- Customer Bill Inputs")

    # write field labels in column B; the values themselves go in C
    # (matching FIELD_TO_CELL).
    for f in FIELDS:
        cell_ref = FIELD_TO_CELL[f.key]   # e.g. "C4"
        # the label sits one column to the left of the input cell
        col_letter = cell_ref[0]
        row = int(cell_ref[1:])
        label_col = chr(ord(col_letter) - 1)
        _label(s1[f"{label_col}{row}"], f.label + (f"  ({f.unit})" if f.unit else ""))
        _input(s1[cell_ref])

    # ---- Sheet 2: Solar calculations -------------------------------------
    s2 = wb.create_sheet("Solar")
    s2.column_dimensions["A"].width = 2
    s2.column_dimensions["B"].width = 38
    s2.column_dimensions["C"].width = 22
    s2.column_dimensions["D"].width = 18

    s2.merge_cells("B2:D2")
    _header(s2["B2"], "Solar Load Calculation")

    rows = [
        ("Avg daily consumption",       f"={INPUT_SHEET}!C7/30",                              "kWh / day"),
        ("Avg peak sun hours (assumed)", 4.5,                                                  "hours"),
        ("Recommended system size",     "=ROUND(C4/C5, 2)",                                   "kW"),
        ("Estimated system cost",       "=C6*55000",                                           "INR"),
        ("Annual generation (yr 1)",    "=C6*C5*365",                                          "kWh / year"),
        ("Avg tariff (assumed)",        9.5,                                                   "INR / kWh"),
        ("Annual savings (yr 1)",       "=C8*C9",                                              "INR / year"),
        ("Payback period",              "=ROUND(C7/C10, 1)",                                   "years"),
        ("25-year net savings",         "=C10*25 - C7",                                        "INR"),
        ("CO2 offset (annual)",         "=ROUND(C8*0.82/1000, 2)",                             "tonnes / year"),
        ("Sanctioned load headroom",    f"=IFERROR({INPUT_SHEET}!C11 - C6, \"--\")",         "kW"),
    ]

    s2["B3"], s2["C3"], s2["D3"] = "Metric", "Value", "Unit"
    for c in (s2["B3"], s2["C3"], s2["D3"]):
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=DARK)
        c.alignment = Alignment(horizontal="left", vertical="center")

    for i, (label, value, unit) in enumerate(rows, start=4):
        s2[f"B{i}"].value = label
        s2[f"B{i}"].font = Font(bold=True, color=DARK)
        s2[f"C{i}"].value = value
        s2[f"D{i}"].value = unit
        # pretty number formatting
        if "INR" in unit:
            s2[f"C{i}"].number_format = "#,##0"
        elif unit == "kW" or unit == "kWh / day":
            s2[f"C{i}"].number_format = "0.00"
        for col in ("B", "C", "D"):
            s2[f"{col}{i}"].border = BORDER
            s2[f"{col}{i}"].fill = PatternFill("solid", fgColor=SOFT_GREY if i % 2 else "FFFFFF")
            s2[f"{col}{i}"].alignment = Alignment(horizontal="left", vertical="center")

    # ---- Sheet 3: Notes (so the explanation lives inside the file) -------
    s3 = wb.create_sheet("Notes")
    s3.column_dimensions["A"].width = 2
    s3.column_dimensions["B"].width = 100

    _header(s3["B2"], "How this sheet works")
    notes = [
        "Inputs sheet holds the values pulled from the customer's bill.",
        "Solar sheet runs the calculations off those inputs -- formulas only,",
        "so re-running the AI on a new bill just refreshes the numbers.",
        "Yellow cells = editable inputs. Everything else is generated.",
        "Assumptions used (sun hours, tariff, cost per kW) live on the Solar sheet",
        "in cells C5, C9, and the cost multiplier in C7's formula -- override per customer.",
    ]
    for i, line in enumerate(notes, start=4):
        s3[f"B{i}"] = line
        s3[f"B{i}"].alignment = Alignment(wrap_text=True, vertical="top")

    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    return out


if __name__ == "__main__":
    p = Path(__file__).resolve().parents[2] / "templates" / "solar_load_template.xlsx"
    build(p)
    print(f"wrote template -> {p}")
