# Offline tests -- no network, no API key required.
# Covers the bits most likely to break:
#   - regex fallback parses obvious MSEDCL fields
#   - excel filler writes inputs but never clobbers a formula
#   - field <-> cell mapping is consistent

import sys
from pathlib import Path

# allow `pytest` from repo root without installing
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT))

import pytest
from openpyxl import load_workbook

from src.config import FIELD_KEYS, FIELD_TO_CELL
from src.excel.filler import fill_template
from src.excel.template_builder import build as build_template
from src.extractor.llm_extractor import extract_from_text_only


SAMPLE_BILL_TEXT = """
MSEDCL Maharashtra State Electricity Distribution Co. Ltd.

Consumer No: 123456789012
Meter No: 9988776655
Billing Period: 01/03/2025 - 31/03/2025
Tariff: LT-I Residential

Units Consumed: 542
Connected Load: 4.5
Sanctioned Load: 5.0

Total Amount Payable: Rs 5,432.50
"""


def test_regex_fallback_picks_up_obvious_fields():
    out = extract_from_text_only(SAMPLE_BILL_TEXT)
    f = out["fields"]
    assert f["consumer_number"] == "123456789012"
    assert f["meter_number"] == "9988776655"
    assert f["units_consumed"] == 542.0
    assert f["total_bill_amount"] == 5432.5
    assert f["connected_load"] == 4.5
    assert f["sanctioned_load"] == 5.0


def test_field_mapping_is_complete():
    # every field defined in config must have a cell
    for key in FIELD_KEYS:
        assert key in FIELD_TO_CELL, f"{key} missing from FIELD_TO_CELL"


def test_filler_writes_inputs_but_keeps_formulas(tmp_path):
    template = tmp_path / "tpl.xlsx"
    build_template(template)

    fields = {
        "consumer_name": "Anurag Test",
        "consumer_number": "123456789012",
        "billing_period": "Mar 2025",
        "units_consumed": 542,
        "total_bill_amount": 5432.5,
        "tariff_category": "LT-I Residential",
        "connected_load": 4.5,
        "sanctioned_load": 5.0,
        "meter_number": "9988776655",
    }

    out = tmp_path / "filled.xlsx"
    fill_template(template, fields, out)

    wb = load_workbook(out, data_only=False)
    inputs = wb["Inputs"]
    solar = wb["Solar"]

    # inputs landed where we expect
    assert inputs["C4"].value == "Anurag Test"
    assert inputs["C7"].value == 542
    assert inputs["C8"].value == 5432.5

    # formulas survived
    assert isinstance(solar["C4"].value, str) and solar["C4"].value.startswith("=")
    assert isinstance(solar["C6"].value, str) and solar["C6"].value.startswith("=")
    assert isinstance(solar["C10"].value, str) and solar["C10"].value.startswith("=")


def test_filler_refuses_to_overwrite_a_formula(tmp_path, monkeypatch):
    # Build a template, then sneak a formula into one of the input cells
    # and confirm the filler refuses to clobber it.
    template = tmp_path / "tpl.xlsx"
    build_template(template)

    wb = load_workbook(template, data_only=False)
    wb["Inputs"]["C7"].value = "=999"   # pretend someone put a formula here
    wb.save(template)

    out = tmp_path / "filled.xlsx"
    fill_template(template, {"units_consumed": 1234}, out)

    wb2 = load_workbook(out, data_only=False)
    assert wb2["Inputs"]["C7"].value == "=999"   # untouched
